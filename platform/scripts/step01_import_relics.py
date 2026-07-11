"""Step 01 | 导入文物数据,生成标准化数据集。

数据源(按优先级):
    A. Markdown 档案  data/input/01_relics/markdown/{分组}/*.md
       (step00 从四普登记表 docx 提取的产物,含简介/DMS坐标/本体边界/清单)
       照片与图纸直接从 data/input/00_docs 对应 docx 内嵌图片按清单顺序抽取。
       若 data/济宁市未清洗数据 存在,再从其中仅追加 rank=5 的未定级文物；
       rank=1..4 的在级文保单位由主档案保持权威,不会重复导入。
    B. 台账 Excel/CSV data/input/01_relics/*.xlsx|csv (旧格式,兼容保留)

两线范围(保护范围/建设控制地带):
    来自外部测绘成果 data/*两线*/ 目录下的 GeoJSON(WGS-84,
    properties: name/county/rangeType 0=保护范围 1=建设控制地带),
    按 名称+县区 匹配台账文物后并入 relics_polygons.geojson。
    (登记表档案中只有本体边界测点,两线不再从档案提取)

输出:
    data/output/dataset/relics_full.json
    data/output/dataset/relics_points.geojson
    data/output/dataset/relics_polygons.geojson   (本体 kind=body + 两线 protection/control)
    data/output/dataset/photo_index.csv / drawing_index.csv
    data/output/photos/{code}/... / data/output/drawings/{code}/...
"""
from __future__ import annotations

import csv
import json
import re
import shutil
import sys
from pathlib import Path

from _common import PROJECT_ROOT, get_logger, get_paths, load_config
from codes import normalize_category, normalize_rank, parse_coord
from md_archive import extract_media_from_docx, parse_archive_md

log = get_logger("step01_import_relics")

# ── 台账中文列名 → 内部字段(Excel 兼容路径用) ─────────────────
_COLUMN_ALIASES: dict[str, str] = {
    "编号": "archive_code", "档案编号": "archive_code", "文物编号": "archive_code", "code": "archive_code",
    "名称": "name", "文物名称": "name", "name": "name",
    "级别": "heritage_level", "保护级别": "heritage_level", "文物级别": "heritage_level",
    "类别": "category_main", "文物类别": "category_main", "分类": "category_main",
    "年代": "era", "时代": "era",
    "年代分期": "era_stats",
    "县区": "county", "县市区": "county", "区县": "county", "所在县区": "county",
    "乡镇": "township", "乡镇街道": "township", "镇街": "township",
    "村": "village", "行政村": "village", "村庄": "village",
    "地址": "address", "详细地址": "address", "位置": "address",
    "经度": "center_lng", "lng": "center_lng",
    "纬度": "center_lat", "lat": "center_lat",
    "高程": "center_alt", "海拔": "center_alt",
    "简介": "intro", "文物简介": "intro", "概况": "intro",
    "保存状况": "condition_level", "保存现状": "condition_level", "现状评估": "condition_level",
    "保护范围": "protection_scope",
    "建设控制地带": "control_zone", "建控地带": "control_zone",
    "数据层级": "tier", "层级": "tier",
    "公布批次": "batch", "批次": "batch",
    "面积": "area",
    "权属": "ownership_type", "产权": "ownership_type",
}

_TIER_ALIASES = {
    "基础": "city", "全市": "city", "city": "city", "市级框架": "city",
    "全量": "full", "嘉祥": "full", "full": "full", "县级全量": "full",
}

_CONDITIONS = ("好", "较好", "一般", "较差", "差")


# ════════════════════════════════════════════════════════════
# A. Markdown 档案导入
# ════════════════════════════════════════════════════════════

def _find_docx_for(md_path: Path, md_root: Path, code: str, docs_root: Path) -> Path | None:
    """按目录镜像/档案号前缀匹配源 docx(照片图纸内嵌其中)。

    step00 输出的 markdown 与 00_docs 目录结构一一镜像,优先同路径同名;
    兜底在同目录按档案号前缀模糊匹配。
    """
    if not docs_root.exists():
        return None
    try:
        rel_dir = md_path.parent.relative_to(md_root)
    except ValueError:
        rel_dir = Path(".")
    group_dir = docs_root / rel_dir
    candidates = [group_dir / f"{md_path.stem}.docx"]
    if group_dir.exists():
        candidates.extend(sorted(group_dir.glob(f"{code}*.docx")))
    for c in candidates:
        if c.exists():
            return c
    return None


def _copy_media_dir(src_root: Path, dst_root: Path, code: str,
                    kinds: tuple[str, ...]) -> list[dict]:
    """input/02_media/{photos|drawings}/{code}/ → output;返回索引行。"""
    index: list[dict] = []
    code_dir = src_root / code
    if not code_dir.exists():
        return index
    for f in sorted(code_dir.rglob("*")):
        if not f.is_file() or f.suffix.lower() not in kinds:
            continue
        rel = f"{code}/{f.name}"
        dst = dst_root / rel
        dst.parent.mkdir(parents=True, exist_ok=True)
        if not dst.exists():
            shutil.copy2(f, dst)
        index.append({"archive_code": code, "path": rel})
    return index


def _supplemental_markdown_root(paths, cfg: dict) -> Path | None:
    """返回“全市未清洗 Markdown”目录。

    默认使用用户约定的 data/济宁市未清洗数据；可通过
    data_import.include_ungraded_markdown=false 关闭，或用
    data_import.ungraded_markdown_dir 指向其他目录。
    """
    data_cfg = cfg.get("data_import") or {}
    if data_cfg.get("include_ungraded_markdown", True) is False:
        return None
    raw = str(data_cfg.get("ungraded_markdown_dir") or "data/济宁市未清洗数据").strip()
    if not raw:
        return None
    root = Path(raw)
    if not root.is_absolute():
        root = paths.root / root
    try:
        if root.resolve() == paths.input_markdown.resolve():
            return None
    except OSError:
        pass
    return root


def _configured_code_set(cfg: dict, key: str) -> set[str]:
    """读取 data_import 下的档案号白名单，兼容单值和 YAML 列表。"""
    raw = (cfg.get("data_import") or {}).get(key) or []
    if isinstance(raw, str):
        raw = [raw]
    return {str(code).strip() for code in raw if str(code).strip()}


def _configured_level_overrides(cfg: dict) -> dict[str, str]:
    """读取受控修复记录的保护级别覆写。"""
    raw = (cfg.get("data_import") or {}).get("protected_repair_level_overrides") or {}
    if not isinstance(raw, dict):
        return {}
    return {
        str(code).strip(): str(level).strip()
        for code, level in raw.items()
        if str(code).strip() and str(level).strip()
    }


def import_from_markdown(paths, cfg: dict) -> tuple[list[dict], list[dict], list[dict], list[dict]] | None:
    """解析 markdown 档案。返回 (relics, photo_index, drawing_index, boundary_features);
    没有 md 文件时返回 None(让调用方走 Excel 路径)。"""
    md_root = paths.input_markdown
    primary_files: list[Path] = []
    if md_root.exists():
        primary_files = sorted(p for p in md_root.rglob("*.md") if not p.name.endswith("_QC.md"))

    supplemental_root = _supplemental_markdown_root(paths, cfg)
    supplemental_files: list[Path] = []
    if supplemental_root and supplemental_root.exists():
        supplemental_files = sorted(
            p for p in supplemental_root.rglob("*.md") if not p.name.endswith("_QC.md")
        )

    if not primary_files and not supplemental_files:
        return None

    log.info("[MD] 主档案 %d 个: %s", len(primary_files), md_root)
    if supplemental_files:
        log.info("[MD] 全市补充档案 %d 个: %s（仅追加未定级）",
                 len(supplemental_files), supplemental_root)

    full_tier_county = ((cfg.get("administrative") or {}).get("full_tier_county") or "").strip()
    src_crs = ((cfg.get("geo") or {}).get("source_crs") or "wgs84").lower()
    protected_repair_codes = _configured_code_set(cfg, "protected_repair_codes")
    invalid_baseline_codes = _configured_code_set(cfg, "invalid_protected_baseline_codes")
    repair_level_overrides = _configured_level_overrides(cfg)

    relics: list[dict] = []
    photo_index: list[dict] = []
    drawing_index: list[dict] = []
    boundary_features: list[dict] = []
    seen: set[str] = set()
    primary_codes: set[str] = set()
    n_fail = n_media_docx = n_media_dir = 0
    report = {
        "primary_discovered": len(primary_files),
        "supplemental_discovered": len(supplemental_files),
        "supplemental_protected_excluded": 0,
        "protected_repair_requested": sorted(protected_repair_codes),
        "protected_repair_accepted": 0,
        "protected_repair_level_overrides": {},
        "primary_invalid_code_excluded": 0,
        "supplemental_existing_code_excluded": 0,
        "supplemental_duplicate_excluded": 0,
        "supplemental_ungraded_accepted": 0,
        "parse_or_validation_failed": 0,
        "rejects": [],
    }

    sources = (
        [(p, md_root, False) for p in primary_files]
        + [(p, supplemental_root, True) for p in supplemental_files]
    )
    for md_path, source_root, is_supplemental in sources:
        assert source_root is not None
        group = md_path.parent.name if md_path.parent != source_root else ""
        try:
            r = parse_archive_md(md_path, group_name=group)
        except Exception as e:  # noqa: BLE001
            log.warning("[MD] 解析失败 %s: %s", md_path.name, e)
            n_fail += 1
            report["parse_or_validation_failed"] += 1
            report["rejects"].append({
                "source": "supplemental" if is_supplemental else "primary",
                "file": str(md_path.relative_to(paths.root)),
                "reason": f"parse_error: {e}",
            })
            continue

        code = str(r.get("archive_code") or "").strip()
        name = str(r.get("name") or "").strip()
        rank_code = normalize_rank(r.get("heritage_level"))

        if not is_supplemental and code in invalid_baseline_codes:
            report["primary_invalid_code_excluded"] += 1
            continue

        is_protected_repair = (
            is_supplemental
            and rank_code in {"1", "2", "3", "4"}
            and code in protected_repair_codes
        )
        if is_supplemental:
            # 补充目录同时含 1,110 处在级文保单位。默认只接收未定级；
            # 已确认的少量历史基线异常可通过精确档案号白名单修复。
            if rank_code != "5" and not is_protected_repair:
                report["supplemental_protected_excluded"] += 1
                continue
            try:
                county_dir = md_path.relative_to(source_root).parts[0]
            except (ValueError, IndexError):
                county_dir = ""
            if re.search(r"[县区市]$", county_dir):
                r["county"] = county_dir
            if is_protected_repair and code in repair_level_overrides:
                level = repair_level_overrides[code]
                if normalize_rank(level) not in {"1", "2", "3", "4"}:
                    raise ValueError(f"白名单修复级别无效: {code} -> {level}")
                r["heritage_level"] = level
                rank_code = normalize_rank(level)
                report["protected_repair_level_overrides"][code] = level

        if not code or not name:
            log.warning("[MD] 缺编号或名称,跳过: %s", md_path.name)
            n_fail += 1
            report["parse_or_validation_failed"] += 1
            report["rejects"].append({
                "source": "supplemental" if is_supplemental else "primary",
                "file": str(md_path.relative_to(paths.root)),
                "code": code,
                "name": name,
                "reason": "missing_code_or_name",
            })
            continue
        if code in seen:
            if is_supplemental:
                if code in primary_codes:
                    report["supplemental_existing_code_excluded"] += 1
                else:
                    report["supplemental_duplicate_excluded"] += 1
            else:
                log.warning("[MD] 重复编号,跳过: %s (%s)", code, md_path.name)
            continue
        if r.get("center_lng") is None or r.get("center_lat") is None:
            log.warning("[MD] 无有效坐标,跳过: %s %s", code, name)
            n_fail += 1
            report["parse_or_validation_failed"] += 1
            report["rejects"].append({
                "source": "supplemental" if is_supplemental else "primary",
                "file": str(md_path.relative_to(paths.root)),
                "code": code,
                "name": name,
                "reason": "missing_valid_coordinate",
            })
            continue
        seen.add(code)
        if not is_supplemental:
            primary_codes.add(code)

        if src_crs == "gcj02":
            from _common import gcj02_to_wgs84
            r["center_lng"], r["center_lat"] = gcj02_to_wgs84(r["center_lng"], r["center_lat"])

        boundary_points = r.pop("_boundary_points", [])
        drawings_meta = r.pop("_drawings_meta", [])
        photos_meta = r.pop("_photos_meta", [])

        # 保存状况兜底 + 层级判定(全量层县 → full)
        cond = r.get("condition_level", "")
        r["condition_level"] = cond if cond in _CONDITIONS else (cond or "一般")
        r["tier"] = "full" if (full_tier_county and full_tier_county in (r.get("county") or "")) else "city"
        r["heritage_level"] = r.get("heritage_level") or "尚未核定公布为文物保护单位的不可移动文物"
        r["category_main"] = r.get("category_main") or "其他"
        if not r.get("era_stats"):
            r["era_stats"] = r.get("era", "")
        r["center_alt"] = r.get("center_alt") if r.get("center_alt") is not None else 0.0
        r["category_code"] = normalize_category(r.get("category_main"))
        r["rank_code"] = normalize_rank(r.get("heritage_level"))
        r["data_scope"] = "ungraded" if r["rank_code"] == "5" else "protected"
        if is_protected_repair:
            r["source_collection"] = "approved_protected_repair"
        elif is_supplemental:
            r["source_collection"] = "supplemental_ungraded"
        else:
            r["source_collection"] = "primary"

        # 档案里的本体边界测点只记数量,不再生成面——
        # 地图上的边界面只用外部两线范围(保护范围/建控地带)。
        # has_boundary 由两线匹配结果决定(见 main)。
        r["has_boundary"] = False
        r["boundary_count"] = len(boundary_points)

        # 媒体:优先从源 docx 抽取内嵌图片,否则回退 02_media 目录
        docx = _find_docx_for(md_path, source_root, code, paths.input_docs)
        if docx and (drawings_meta or photos_meta):
            try:
                d_rows, p_rows = extract_media_from_docx(
                    docx, code, drawings_meta, photos_meta,
                    paths.output_drawings, paths.output_photos,
                )
                drawing_index.extend(d_rows)
                photo_index.extend(p_rows)
                n_media_docx += 1
            except Exception as e:  # noqa: BLE001
                log.warning("[MD] %s 抽取 docx 图片失败: %s", code, e)
        else:
            p_rows = _copy_media_dir(paths.input_media / "photos", paths.output_photos,
                                     code, (".jpg", ".jpeg", ".png", ".webp"))
            d_rows = _copy_media_dir(paths.input_media / "drawings", paths.output_drawings,
                                     code, (".jpg", ".jpeg", ".png", ".webp", ".pdf"))
            photo_index.extend(p_rows)
            drawing_index.extend(d_rows)
            if p_rows or d_rows:
                n_media_dir += 1

        relics.append(r)
        if is_protected_repair:
            report["protected_repair_accepted"] += 1
        elif is_supplemental:
            report["supplemental_ungraded_accepted"] += 1

    log.info("[MD] 解析成功 %d / 失败 %d;媒体来源: docx %d 处 / 目录 %d 处",
             len(relics), n_fail, n_media_docx, n_media_dir)
    if supplemental_files:
        log.info("[MD] 补充数据筛选: 排除在级 %d / 白名单修复 %d / 编号冲突 %d / "
                 "重复编号 %d / 接收未定级 %d",
                 report["supplemental_protected_excluded"],
                 report["protected_repair_accepted"],
                 report["supplemental_existing_code_excluded"],
                 report["supplemental_duplicate_excluded"],
                 report["supplemental_ungraded_accepted"])
        paths.output_dataset.mkdir(parents=True, exist_ok=True)
        (paths.output_dataset / "ungraded_import_report.json").write_text(
            json.dumps(report, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    return relics, photo_index, drawing_index, boundary_features


# ════════════════════════════════════════════════════════════
# B. Excel/CSV 台账导入(旧格式兼容)
# ════════════════════════════════════════════════════════════

def _find_source(input_dir: Path) -> Path | None:
    for pat in ("*.xlsx", "*.xls", "*.csv"):
        files = sorted(p for p in input_dir.glob(pat) if not p.name.startswith("~"))
        if files:
            return files[0]
    return None


def _read_rows(src: Path) -> list[dict]:
    if src.suffix.lower() == ".csv":
        with src.open("r", encoding="utf-8-sig") as f:
            return list(csv.DictReader(f))

    try:
        from openpyxl import load_workbook
    except ImportError:
        log.error("未安装 openpyxl,无法读取 Excel。请 pip install openpyxl,或改用 CSV。")
        sys.exit(1)

    wb = load_workbook(src, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        return []
    out = []
    for values in rows_iter:
        if values is None or all(v in (None, "") for v in values):
            continue
        out.append({header[i]: values[i] for i in range(min(len(header), len(values)))})
    wb.close()
    return out


def _normalize_row(raw: dict) -> dict:
    r: dict = {}
    for k, v in raw.items():
        key = _COLUMN_ALIASES.get(str(k).strip())
        if not key:
            continue
        if v is None:
            continue
        s = str(v).strip()
        if s:
            r[key] = s

    code = r.get("archive_code", "")
    name = r.get("name", "")
    lng = parse_coord(r.get("center_lng"))
    lat = parse_coord(r.get("center_lat"))
    if not code or not name or lng is None or lat is None:
        return {}

    r["archive_code"] = code
    r["center_lng"] = lng
    r["center_lat"] = lat
    alt = parse_coord(r.get("center_alt"))
    r["center_alt"] = alt if alt is not None else 0.0

    r["heritage_level"] = r.get("heritage_level") or "未定级"
    r["category_main"] = r.get("category_main") or "其他"
    cond = r.get("condition_level", "")
    r["condition_level"] = cond if cond in _CONDITIONS else (cond or "一般")

    tier_raw = r.get("tier", "")
    r["tier"] = _TIER_ALIASES.get(tier_raw, "full" if tier_raw == "full" else "city")

    if not r.get("era_stats"):
        r["era_stats"] = r.get("era", "")
    return r


def _copy_media(src_root: Path, dst_root: Path, kinds: tuple[str, ...]) -> list[dict]:
    index: list[dict] = []
    if not src_root.exists():
        return index
    for code_dir in sorted(src_root.iterdir()):
        if not code_dir.is_dir():
            continue
        code = code_dir.name
        for f in sorted(code_dir.rglob("*")):
            if not f.is_file() or f.suffix.lower() not in kinds:
                continue
            rel = f"{code}/{f.name}"
            dst = dst_root / rel
            dst.parent.mkdir(parents=True, exist_ok=True)
            if not dst.exists():
                shutil.copy2(f, dst)
            index.append({"archive_code": code, "path": rel})
    return index


def import_from_excel(paths, cfg: dict) -> tuple[list[dict], list[dict], list[dict], list[dict]] | None:
    src = _find_source(paths.input_relics)
    if not src:
        return None
    src_crs = ((cfg.get("geo") or {}).get("source_crs") or "wgs84").lower()

    log.info("[Excel] 台账: %s", src.name)
    raw_rows = _read_rows(src)
    log.info("[Excel] 读到 %d 行", len(raw_rows))

    relics: list[dict] = []
    seen: set[str] = set()
    for raw in raw_rows:
        r = _normalize_row(raw)
        if not r:
            continue
        if r["archive_code"] in seen:
            log.warning("重复编号,跳过: %s", r["archive_code"])
            continue
        seen.add(r["archive_code"])
        if src_crs == "gcj02":
            from _common import gcj02_to_wgs84
            r["center_lng"], r["center_lat"] = gcj02_to_wgs84(r["center_lng"], r["center_lat"])
        r["category_code"] = normalize_category(r.get("category_main"))
        r["rank_code"] = normalize_rank(r.get("heritage_level"))
        relics.append(r)

    if not relics:
        log.error("有效记录为 0,请检查台账列名与内容")
        return None

    photo_index = _copy_media(paths.input_media / "photos", paths.output_photos,
                              (".jpg", ".jpeg", ".png", ".webp"))
    drawing_index = _copy_media(paths.input_media / "drawings", paths.output_drawings,
                                (".jpg", ".jpeg", ".png", ".webp", ".pdf"))
    return relics, photo_index, drawing_index, []


# ════════════════════════════════════════════════════════════
# 公共产物输出
# ════════════════════════════════════════════════════════════

def _scan_archive_docs(root: Path) -> dict[str, dict[str, int]]:
    out: dict[str, dict[str, int]] = {}
    if not root.exists():
        return out
    for code_dir in root.iterdir():
        if not code_dir.is_dir():
            continue
        entry = {"sanpu": 0, "sipu": 0}
        for kind in ("sanpu", "sipu"):
            d = code_dir / kind
            if d.exists():
                entry[kind] = sum(1 for p in d.glob("*.pdf")) + sum(1 for p in d.glob("*.PDF"))
        if entry["sanpu"] or entry["sipu"]:
            out[code_dir.name] = entry
    return out


def _write_points_geojson(relics: list[dict], out: Path) -> None:
    feats = []
    for r in relics:
        feats.append({
            "type": "Feature",
            "geometry": {"type": "Point", "coordinates": [r["center_lng"], r["center_lat"]]},
            "properties": {
                "archive_code": r["archive_code"],
                "name": r["name"],
                "heritage_level": r.get("heritage_level", ""),
                "county": r.get("county", ""),
            },
        })
    out.write_text(
        json.dumps({"type": "FeatureCollection", "features": feats}, ensure_ascii=False),
        encoding="utf-8",
    )


# ── 市保/县保级别精化 ────────────────────────────────────────
# 四普登记表把市保与县保合并为一个勾选项「市级和县级文物保护单位」。
# 用 data/ 下的两份权威名单(市保名单 xls + 县保实有名录 xlsx)拆分,
# 名单未覆盖的再看简介中的公布语句,最后兜底县保。
_MERGED_LEVEL = "市级和县级文物保护单位"
_CITY_LEVEL = "市级文物保护单位"
_COUNTY_LEVEL = "县级文物保护单位"


def _norm_name(s) -> str:
    import re as _re
    t = _re.sub(r"\s+", "", str(s or ""))
    return t.replace("（", "(").replace("）", ")")


def _load_rank_rosters() -> tuple[dict, dict] | None:
    """读市保/县保名单。返回 (city_idx, county_idx),
    索引结构 {norm_name: {county...}};任一名单缺失返回 None(跳过精化)。"""
    data_dir = PROJECT_ROOT / "data"
    if not data_dir.exists():
        return None
    xlsx = next((p for p in data_dir.glob("*.xlsx") if "县级" in p.name and "名录" in p.name), None)
    xls = next((p for p in data_dir.glob("*.xls")
                if p.suffix.lower() == ".xls" and "名单" in p.name), None)
    if not xlsx or not xls:
        return None

    county_idx: dict[str, set[str]] = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=2, max_col=10, values_only=True):
                name = row[3] if len(row) > 3 else None
                county = row[7] if len(row) > 7 else ""
                if not name or str(name).strip() in ("文物保护单位名称",):
                    continue
                county_idx.setdefault(_norm_name(name), set()).add(str(county or "").strip())
        wb.close()
    except Exception as e:  # noqa: BLE001
        log.warning("[级别] 县保名录读取失败(%s): %s", xlsx.name, e)
        return None

    city_idx: dict[str, set[str]] = {}
    try:
        import xlrd
        wb2 = xlrd.open_workbook(str(xls))
        for sheet in wb2.sheets():
            if "济宁市文物保护单位" not in sheet.name and "市文物保护单位" not in sheet.name:
                continue
            for r in range(sheet.nrows):
                vals = sheet.row_values(r)
                if len(vals) < 5 or not vals[1]:
                    continue
                try:
                    float(vals[0])  # 数据行序号为数字
                except (TypeError, ValueError):
                    continue
                city_idx.setdefault(_norm_name(vals[1]), set()).add(str(vals[4] or "").strip())
    except ImportError:
        log.warning("[级别] 未安装 xlrd,无法读取市保名单(pip install xlrd)")
        return None
    except Exception as e:  # noqa: BLE001
        log.warning("[级别] 市保名单读取失败(%s): %s", xls.name, e)
        return None

    log.info("[级别] 名单载入: 市保 %d 处 / 县保 %d 处", len(city_idx), len(county_idx))
    return city_idx, county_idx


def _roster_hit(idx: dict[str, set[str]], name: str, county: str) -> bool:
    """名单命中判断:名称精确 + 县区一致(名单地址含县区名或未填县区)。"""
    entry = idx.get(name)
    if entry is None:
        return False
    if not county:
        return True
    return any((not loc) or (county in loc) or (loc in county) for loc in entry)


def _fuzzy_hit(idx: dict[str, set[str]], name: str) -> bool:
    """模糊命中:名称包含关系(限 >=4 字,避免「张氏家祠」这类泛称误配)。"""
    if len(name) < 4:
        return False
    return any(len(k) >= 4 and (k in name or name in k) for k in idx)


def _refine_merged_levels(relics: list[dict]) -> None:
    """把 heritage_level=「市级和县级文物保护单位」拆分为市保/县保(就地覆写)。"""
    import re as _re

    merged = [r for r in relics if (r.get("heritage_level") or "").strip() == _MERGED_LEVEL]
    if not merged:
        return
    rosters = _load_rank_rosters()
    if rosters is None:
        log.warning("[级别] 未找到市/县保名单 Excel,%d 处「市级和县级」保持市保口径", len(merged))
        return
    city_idx, county_idx = rosters

    re_cnty = _re.compile(r"(?:公布|核定|批准|定)[^。;；]{0,25}[县区]级(?:重点)?文物保护单位")
    re_city = _re.compile(r"(?:公布|核定|批准|定)[^。;；]{0,25}市级(?:重点)?文物保护单位")

    stat = {"roster_city": 0, "roster_county": 0, "intro_city": 0, "intro_county": 0,
            "fuzzy_city": 0, "fuzzy_county": 0, "default_county": 0}

    def set_level(r: dict, level: str) -> None:
        r["heritage_level"] = level
        r["rank_code"] = normalize_rank(level)

    for r in merged:
        name = _norm_name(r.get("name"))
        county = (r.get("county") or "").strip()
        in_city = _roster_hit(city_idx, name, county)
        in_cnty = _roster_hit(county_idx, name, county)
        if in_city and in_cnty:
            in_cnty = False  # 两名单同名同县(极少):就高按市保
        if in_city:
            set_level(r, _CITY_LEVEL); stat["roster_city"] += 1
            continue
        if in_cnty:
            set_level(r, _COUNTY_LEVEL); stat["roster_county"] += 1
            continue
        brief = r.get("intro") or ""
        has_c, has_s = bool(re_cnty.search(brief)), bool(re_city.search(brief))
        if has_s and not has_c:
            set_level(r, _CITY_LEVEL); stat["intro_city"] += 1
            continue
        if has_c and not has_s:
            set_level(r, _COUNTY_LEVEL); stat["intro_county"] += 1
            continue
        if _fuzzy_hit(city_idx, name):
            set_level(r, _CITY_LEVEL); stat["fuzzy_city"] += 1
            continue
        if _fuzzy_hit(county_idx, name):
            set_level(r, _COUNTY_LEVEL); stat["fuzzy_county"] += 1
            continue
        # 无任何证据:市保名单是封闭名录(198 处)且已尽数命中,兜底县保
        set_level(r, _COUNTY_LEVEL); stat["default_county"] += 1

    n_city = stat["roster_city"] + stat["intro_city"] + stat["fuzzy_city"]
    n_cnty = len(merged) - n_city
    log.info("[级别] 市县保拆分: 共 %d 处 → 市保 %d / 县保 %d "
             "(名单 %d+%d, 简介 %d+%d, 模糊 %d+%d, 兜底县保 %d)",
             len(merged), n_city, n_cnty,
             stat["roster_city"], stat["roster_county"],
             stat["intro_city"], stat["intro_county"],
             stat["fuzzy_city"], stat["fuzzy_county"], stat["default_county"])


# ── 外部两线范围导入 ─────────────────────────────────────────
# rangeType: 0=保护范围(protection) 1=建设控制地带(control)
_RANGE_TYPE_KIND = {"0": "protection", "1": "control"}


def _find_two_line_dirs() -> list[Path]:
    """定位两线范围数据目录:data/ 下目录名含「两线」的全部目录。"""
    data_dir = PROJECT_ROOT / "data"
    if not data_dir.exists():
        return []
    return sorted(d for d in data_dir.iterdir() if d.is_dir() and "两线" in d.name)


def _load_two_line_features() -> list[dict]:
    """读取两线目录下全部 GeoJSON,按 polyId 去重(分县文件与汇总文件重叠)。
    优先分县文件(文件名 城市_县区.geojson);没有分县文件才读汇总。"""
    dirs = _find_two_line_dirs()
    if not dirs:
        return []
    by_id: dict[str, dict] = {}
    for d in dirs:
        files = sorted(d.glob("*_*.geojson")) or sorted(d.glob("*.geojson"))
        for p in files:
            try:
                gj = json.loads(p.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError, UnicodeDecodeError) as e:
                log.warning("[两线] 读取失败,跳过 %s: %s", p.name, e)
                continue
            for f in gj.get("features") or []:
                props = f.get("properties") or {}
                pid = str(props.get("polyId") or f"{p.name}:{len(by_id)}")
                by_id.setdefault(pid, f)
        if by_id:
            log.info("[两线] 数据目录 %s: %d 个面(按 polyId 去重)", d.name, len(by_id))
    return list(by_id.values())


def _match_two_lines(relics: list[dict]) -> list[dict]:
    """把外部两线面匹配到台账文物,返回带 archive_code/kind 的 Feature 列表。

    保护范围/建设控制地带只属于 rank 1..4 的在级文保单位；未定级点即使
    与两线数据同名，也不得吸收这些面。

    匹配策略(依次):
    1. 名称+县区 精确匹配
    2. 名称唯一时仅按名称
    未匹配的记录数量并抽样告警,不阻断导入。
    """
    feats = _load_two_line_features()
    if not feats:
        return []

    by_name_county: dict[tuple[str, str], str] = {}
    by_name: dict[str, list[str]] = {}
    for r in relics:
        if normalize_rank(r.get("heritage_level")) not in {"1", "2", "3", "4"}:
            continue
        name = (r.get("name") or "").strip()
        county = (r.get("county") or "").strip()
        code = r.get("archive_code") or ""
        if not name or not code:
            continue
        by_name_county.setdefault((name, county), code)
        by_name.setdefault(name, []).append(code)

    out: list[dict] = []
    unmatched: dict[str, int] = {}
    kind_count = {"protection": 0, "control": 0}
    for f in feats:
        props = f.get("properties") or {}
        name = (props.get("name") or "").strip()
        county = (props.get("county") or "").strip()
        kind = _RANGE_TYPE_KIND.get(str(props.get("rangeType")))
        if not name or kind is None:
            continue
        geom = f.get("geometry")
        if not geom or geom.get("type") not in ("Polygon", "MultiPolygon"):
            continue
        code = by_name_county.get((name, county))
        if not code:
            codes = by_name.get(name) or []
            if len(set(codes)) == 1:
                code = codes[0]
        if not code:
            unmatched[name] = unmatched.get(name, 0) + 1
            continue
        kind_count[kind] += 1
        out.append({
            "type": "Feature",
            "properties": {"archive_code": code, "kind": kind, "name": name},
            "geometry": geom,
        })

    log.info("[两线] 匹配成功 %d 个面(保护范围 %d / 建控地带 %d)",
             len(out), kind_count["protection"], kind_count["control"])
    if unmatched:
        sample = list(unmatched)[:8]
        log.warning("[两线] %d 处文物未匹配到台账(共 %d 个面),样例: %s%s",
                    len(unmatched), sum(unmatched.values()), "、".join(sample),
                    " …" if len(unmatched) > 8 else "")
    return out


def _write_polygons_geojson(boundary_features: list[dict], paths, out: Path,
                            two_line_features: list[dict] | None = None) -> None:
    """本体边界(md 解析) + 外部两线范围面合并输出。
    也兼容旧的 data/input/01_relics/protection_zones.geojson。"""
    feats = list(boundary_features)
    if two_line_features:
        feats.extend(two_line_features)
    zones = paths.input_relics / "protection_zones.geojson"
    if zones.exists():
        try:
            data = json.loads(zones.read_text(encoding="utf-8"))
            extra = data.get("features") or []
            feats.extend(extra)
            log.info("[导出] 合并 protection_zones.geojson %d 个要素", len(extra))
        except Exception as e:  # noqa: BLE001
            log.warning("[导出] protection_zones.geojson 解析失败: %s", e)
    if not feats:
        return
    out.write_text(
        json.dumps({"type": "FeatureCollection", "features": feats}, ensure_ascii=False),
        encoding="utf-8",
    )
    log.info("[导出] relics_polygons.geojson: %d 个面", len(feats))


def _write_index_csv(rows: list[dict], out: Path, fields: list[str]) -> None:
    with out.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def _load_protected_baseline(path: Path) -> list[dict]:
    """读取现有标准数据集中的 rank 1..4，作为在级文保单位权威基线。

    新下载目录内的 1,110 份在级 Markdown 仅用于识别并排除，绝不覆盖
    已入库的编号、坐标、级别、两线和资源标记。
    """
    if not path.exists():
        return []
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        log.warning("[基线] 现有 relics_full.json 读取失败: %s", exc)
        return []
    if not isinstance(payload, list):
        return []
    protected = [
        dict(r) for r in payload
        if isinstance(r, dict) and normalize_rank(r.get("heritage_level")) in {"1", "2", "3", "4"}
    ]
    return protected


def _merge_protected_baseline(
    baseline: list[dict],
    repair_records: list[dict],
    repair_codes: set[str],
    invalid_codes: set[str],
) -> tuple[list[dict], dict]:
    """以精确白名单修复在级基线，并返回可写入导入报告的审计信息。"""
    repairs: dict[str, dict] = {}
    for row in repair_records:
        code = str(row.get("archive_code") or "").strip()
        if code not in repair_codes:
            continue
        if normalize_rank(row.get("heritage_level")) not in {"1", "2", "3", "4"}:
            raise ValueError(f"白名单修复记录不是在级文物: {code}")
        if code in repairs:
            raise ValueError(f"白名单修复记录编号重复: {code}")
        repairs[code] = row

    missing = sorted(repair_codes - set(repairs))
    if missing:
        raise ValueError(f"白名单修复记录缺失或解析失败: {', '.join(missing)}")

    baseline_codes = {str(row.get("archive_code") or "").strip() for row in baseline}
    removed_invalid = sorted(code for code in baseline_codes if code in invalid_codes)
    replaced_existing = sorted(code for code in baseline_codes if code in repair_codes)
    preserved = [
        row for row in baseline
        if str(row.get("archive_code") or "").strip() not in invalid_codes | repair_codes
    ]
    merged = preserved + [repairs[code] for code in sorted(repairs)]

    merged_codes = [str(row.get("archive_code") or "").strip() for row in merged]
    if not all(merged_codes) or len(set(merged_codes)) != len(merged_codes):
        raise ValueError("修复后的在级基线含空编号或重复编号")

    audit = {
        "protected_baseline_input": len(baseline),
        "protected_baseline_preserved": len(preserved),
        "protected_baseline_invalid_removed": removed_invalid,
        "protected_baseline_replaced_codes": replaced_existing,
        "protected_repair_codes": sorted(repairs),
        "protected_final_total": len(merged),
    }
    return merged, audit


def main() -> int:
    paths = get_paths()
    cfg = load_config()
    existing_dataset = paths.output_dataset / "relics_full.json"
    protected_baseline = _load_protected_baseline(existing_dataset)

    result = import_from_markdown(paths, cfg)
    source = "markdown"
    if result is None:
        result = import_from_excel(paths, cfg)
        source = "excel"
    if result is None:
        if (paths.output_dataset / "relics_full.json").exists():
            log.info("未找到 markdown 档案或台账,但 relics_full.json 已存在,跳过导入。")
            return 0
        log.error("未找到数据源。请把 markdown 档案放入 %s,或台账 Excel/CSV 放入 %s。",
                  paths.input_markdown, paths.input_relics)
        return 2

    relics, photo_index, drawing_index, boundary_features = result
    supplemental_root = _supplemental_markdown_root(paths, cfg)
    supplemental_enabled = bool(
        supplemental_root
        and supplemental_root.exists()
        and next(supplemental_root.rglob("*.md"), None)
    )
    supplemental_ungraded = [
        r for r in relics
        if r.get("source_collection") == "supplemental_ungraded"
        and normalize_rank(r.get("heritage_level")) == "5"
    ]
    protected_repairs = [
        r for r in relics
        if r.get("source_collection") == "approved_protected_repair"
    ]
    if supplemental_enabled and protected_baseline:
        repair_codes = _configured_code_set(cfg, "protected_repair_codes")
        invalid_codes = _configured_code_set(cfg, "invalid_protected_baseline_codes")
        try:
            protected_baseline, baseline_audit = _merge_protected_baseline(
                protected_baseline,
                protected_repairs,
                repair_codes,
                invalid_codes,
            )
        except ValueError as exc:
            log.error("[基线] 受控修复失败，保留现有产物不写出: %s", exc)
            return 3
        for r in protected_baseline:
            r["data_scope"] = "protected"
            if r.get("source_collection") != "approved_protected_repair":
                r["source_collection"] = "protected_baseline"
        relics = protected_baseline + supplemental_ungraded
        source = "protected_baseline+approved_repairs+supplemental_ungraded"
        log.info("[基线] 原样保留在级 %d 条 / 移除异常 %d 条 / 白名单修复 %d 条 / "
                 "最终在级 %d 条 / 追加未定级 %d 条",
                 baseline_audit["protected_baseline_preserved"],
                 len(baseline_audit["protected_baseline_invalid_removed"]),
                 len(baseline_audit["protected_repair_codes"]),
                 len(protected_baseline), len(supplemental_ungraded))
        report_path = paths.output_dataset / "ungraded_import_report.json"
        try:
            report = json.loads(report_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            report = {}
        report.update({
            **baseline_audit,
            "final_ungraded_total": len(supplemental_ungraded),
            "final_all_total": len(relics),
            "protected_source_policy": (
                "现有 relics_full.json 为权威基线；补充目录中的 rank 1..4 默认排除，"
                "仅精确档案号白名单可替换已确认异常"
            ),
        })
        report_path.write_text(
            json.dumps(report, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    # 普查档案 PDF / 三维模型标记 + 媒体计数
    archives = _scan_archive_docs(paths.input_archive_docs)
    models_root = paths.input_models_3d

    valid_codes = {str(r.get("archive_code") or "") for r in relics}
    photo_index = [p for p in photo_index if p.get("archive_code") in valid_codes]
    drawing_index = [d for d in drawing_index if d.get("archive_code") in valid_codes]

    photo_count: dict[str, int] = {}
    for p in photo_index:
        photo_count[p["archive_code"]] = photo_count.get(p["archive_code"], 0) + 1
    drawing_count: dict[str, int] = {}
    for d in drawing_index:
        drawing_count[d["archive_code"]] = drawing_count.get(d["archive_code"], 0) + 1

    # 市保/县保拆分(登记表合并勾选项按权威名单+简介精化)
    _refine_merged_levels(relics)

    # 外部两线范围(名称+县区匹配);匹配到的文物打 has_boundary 标记
    two_line_features = _match_two_lines(relics)
    two_line_codes = {f["properties"]["archive_code"] for f in two_line_features}

    for r in relics:
        code = r["archive_code"]
        r["photo_count"] = photo_count.get(code, 0)
        r["drawing_count"] = drawing_count.get(code, 0)
        arch = archives.get(code, {})
        r["has_archive_spu"] = bool(arch.get("sanpu"))
        r["has_archive_fpu"] = bool(arch.get("sipu"))
        model_dir = models_root / code
        r["has_3d"] = model_dir.exists() and any(model_dir.iterdir()) if model_dir.exists() else bool(r.get("has_3d"))
        if code in two_line_codes:
            r["has_boundary"] = True

    paths.output_dataset.mkdir(parents=True, exist_ok=True)
    out_json = paths.output_dataset / "relics_full.json"
    out_json.write_text(json.dumps(relics, ensure_ascii=False, indent=1), encoding="utf-8")
    log.info("[导出] relics_full.json: %d 条 (数据源: %s)", len(relics), source)

    _write_points_geojson(relics, paths.output_dataset / "relics_points.geojson")
    _write_polygons_geojson(boundary_features, paths, paths.output_dataset / "relics_polygons.geojson",
                            two_line_features=two_line_features)

    _write_index_csv(photo_index, paths.output_dataset / "photo_index.csv",
                     ["archive_code", "path", "photo_no", "description"])
    _write_index_csv(drawing_index, paths.output_dataset / "drawing_index.csv",
                     ["archive_code", "path", "drawing_no", "drawing_name"])
    log.info("[导出] 照片 %d 张 / 图纸 %d 张", len(photo_index), len(drawing_index))

    by_tier = {"city": 0, "full": 0}
    for r in relics:
        by_tier[r.get("tier", "city")] = by_tier.get(r.get("tier", "city"), 0) + 1
    log.info("[导出] 完成。基础层 %d 条 / 全量层 %d 条", by_tier.get("city", 0), by_tier.get("full", 0))
    return 0


if __name__ == "__main__":
    sys.exit(main())
